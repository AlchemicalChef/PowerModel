#!/usr/bin/env python3
"""Extract the EIA-860 sheets the generator ingest reads from the published zip.

`PowerModel.Ingestion.EIA.Form860.ingest/1` reads two CSVs whose first row must
be the column headers:

    data/3_1_Generator_Y<year>.csv   (Schedule 3.1, "Operable" sheet)
    data/2___Plant_Y<year>.csv       (Schedule 2, the only source of coordinates)

EIA ships only `data/eia860_<year>.zip` of XLSX workbooks, each carrying a
one-line title row ("2024 Form EIA-860 Data") above the headers.  This script
unpacks the two workbooks, converts the right sheet of each to CSV, and drops
that title row.  Same openpyxl pattern as
`PowerModel.Ingestion.EPA.EGrid.extract_sheet_to_csv/2`.

Dependencies: openpyxl (`python3 -m pip install openpyxl`).

Usage:

    python3 scripts/prepare_eia860.py                 # data/eia860_2024.zip -> data/
    python3 scripts/prepare_eia860.py --data-dir data --year 2024
    python3 scripts/prepare_eia860.py --force         # rewrite existing CSVs

`mix power_model.ingest prepare_eia860` runs this, and the generators stage of
`full_pipeline` runs it automatically when the CSVs are missing.
"""

from __future__ import annotations

import argparse
import csv
import sys
import tempfile
import zipfile
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent

# (member in the zip, sheet to export, output csv name)
SHEETS = [
    ("3_1_Generator_Y{year}.xlsx", "Operable", "3_1_Generator_Y{year}.csv"),
    ("2___Plant_Y{year}.xlsx", "Plant", "2___Plant_Y{year}.csv"),
]

# The row above the headers in every published EIA-860 sheet.
TITLE_ROW_PREFIX = "Form EIA-860"


def sheet_to_csv(xlsx_path: Path, sheet_name: str, out_path: Path) -> int:
    import openpyxl

    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheet = None
        for name in workbook.sheetnames:
            if name.strip().lower() == sheet_name.strip().lower():
                sheet = workbook[name]
                break
        if sheet is None:
            raise SystemExit(
                f"ERROR: sheet {sheet_name!r} not in {xlsx_path.name} "
                f"(has {workbook.sheetnames})"
            )

        rows_written = 0
        with out_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            for index, row in enumerate(sheet.iter_rows(values_only=True)):
                values = ["" if cell is None else str(cell) for cell in row]
                # Drop the title row so the header row lands first, which is
                # what every column lookup in form860.ex assumes.
                if index == 0 and TITLE_ROW_PREFIX in (values[0] if values else ""):
                    continue
                if not any(value.strip() for value in values):
                    continue
                writer.writerow(values)
                rows_written += 1
        return rows_written
    finally:
        workbook.close()


def prepare(data_dir: Path, year: int, force: bool) -> int:
    zip_path = data_dir / f"eia860_{year}.zip"
    if not zip_path.exists():
        raise SystemExit(
            f"ERROR: {zip_path} not found. Download the EIA-860 annual zip from "
            "https://www.eia.gov/electricity/data/eia860/ into data/."
        )

    written = 0
    with zipfile.ZipFile(zip_path) as archive:
        members = set(archive.namelist())
        with tempfile.TemporaryDirectory() as tmp:
            for member_pattern, sheet, out_pattern in SHEETS:
                member = member_pattern.format(year=year)
                out_path = data_dir / out_pattern.format(year=year)

                if out_path.exists() and not force:
                    print(f"  {out_path.name} already present (--force to rewrite)")
                    continue

                if member not in members:
                    raise SystemExit(f"ERROR: {member} not in {zip_path} (has {sorted(members)})")

                extracted = Path(archive.extract(member, tmp))
                rows = sheet_to_csv(extracted, sheet, out_path)
                print(f"  {out_path.name}: {rows} rows (header + data) from sheet {sheet!r}")
                written += 1

    return written


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--data-dir", type=Path, default=REPO_ROOT / "data")
    parser.add_argument("--year", type=int, default=2024)
    parser.add_argument("--force", action="store_true", help="rewrite CSVs that already exist")
    args = parser.parse_args(argv)

    print(f"Preparing EIA-860 {args.year} CSVs in {args.data_dir}")
    prepare(args.data_dir, args.year, args.force)
    print("Done.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
